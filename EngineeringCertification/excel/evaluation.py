from enum import Enum
import openpyxl as pyxls
import pandas as pd
from pandas.core.frame import DataFrame
from shutil import copyfile

from config import *
from utils.error import *
from utils.excel import *


# 定义枚举类型，分别表示4个年级
class EGrads(Enum):
    GRAD_2014 = 2014
    GRAD_2015 = 2015
    GRAD_2016 = 2016
    GRAD_2017 = 2017


class ExcelBuilder:
    # 输出目录
    out_path = out_path + "\\excel"
    # 数据目录
    data_path = data_path + "\\excel"
    # 保存所有成绩excel数据
    gdfAllData: dict[EGrads:DataFrame] = {EGrads(2014): DataFrame(),
                                          EGrads(2015): DataFrame(),
                                          EGrads(2016): DataFrame(),
                                          EGrads(2017): DataFrame()}
    # 保存所有课程excel数据
    gdfCoursePriority: DataFrame = DataFrame()

    # 预读取所有成绩数据
    for grad in EGrads:
        gdfAllData[grad] = pd.DataFrame(
            pd.read_excel(data_path + f"\\score-{grad.value}.xlsx", sheet_name="Sheet0"))
    # 预读取所有课程数据
    gdfCoursePriority = pd.DataFrame(
        pd.read_excel(data_path + "/classes.xlsx", sheet_name="Sheet0"))

    # 构造函数
    def __init__(self, g_student_grade: EGrads = EGrads(2017)):
        self.gStudentGrade: EGrads = g_student_grade  # 要操作的年级
        self.dfCoursePriority = None  # 课程优先级数据，DataFrame
        self.dfAllData = None  # 整个年级的数据，DataFrame

        self.load_data()

    def set_garde(self, grade: EGrads):
        self.gStudentGrade = grade
        self.load_data()

    def load_data(self):
        """加载数据"""
        self.dfAllData = self.gdfAllData[self.gStudentGrade]
        self.dfCoursePriority = self.gdfCoursePriority

    def modify_excel_src(self):
        """
        修改源 Excel 文件，增加达成度一列，根据成绩进行折算
        只要跑一次就可以。
        """
        # 生成 excel 数据文件地址
        xls_filename = os.path.join(self.data_path,
                                    f"\\score-{self.gStudentGrade.value}.xlsx")

        if not os.path.exists(xls_filename):
            # 数据不存在，抛出异常
            raise DataNotExistError(f"({xls_filename}) does not exist")
        # 打开刚才创建的 workbook
        wb = pyxls.load_workbook(xls_filename)

        # 指定当前显示（活动）的sheet对象
        ws = wb.active

        for tRowIndex in range(2, ws.max_row + 1):
            print(tRowIndex)

            id_modified_grade = "X{}".format(tRowIndex)
            grade_value = ws.cell(row=tRowIndex, column=15).value

            if "优" == grade_value:
                ws[id_modified_grade] = 0.95
                print("优")
            elif "良" == grade_value:
                ws[id_modified_grade] = 0.85
                print("良")
            elif "中" == grade_value:
                ws[id_modified_grade] = 0.75
                print("中")
            elif "及格" == grade_value:
                ws[id_modified_grade] = 0.65
                print("及格")
            elif "不及格" == grade_value:
                ws[id_modified_grade] = 0
                print("不及格")
            else:
                ws[id_modified_grade] = float(grade_value) / 100.0
                print("数值成绩")

        # 保存excel
        wb.save(xls_filename)

    def get_course_priority(self, course_id_in) -> int:
        """
        获取某门课程的排序优先级
        :param course_id_in: 课程编号
        :return: 排序优先级
        """
        try:
            one_course = self.dfCoursePriority[
                self.dfCoursePriority["课程号"] == course_id_in]
            priority_ret = int(one_course["编号"])  # 若查询结果不存在则抛出 NameError
        except:
            # 课程编号不存在的情况下直接给一个大的数值，表示低优先级
            priority_ret = 999

        return priority_ret

    def get_student_ids(self):
        """
        获取某个年级所有学生的学号
        :return: 以集合的形式返回
        """
        stu_dup_ids = self.dfAllData['学号']
        stu_uniq_ids = set()
        for tId in stu_dup_ids:
            stu_uniq_ids.add(tId)
        return stu_uniq_ids

    def get_one_student_info(self, stud_id_in) -> DataFrame:
        """
        获取某个年级的某个学生的所有记录
        :param stud_id_in: 输入的学生编号
        :return: 学生信息的 DataFrame
        """
        # todo 改成用数据库的
        df_one_student_all = self.dfAllData[self.dfAllData['学号'] == stud_id_in]

        # 对数据进行筛选，删除其中绩点为 0 的记录（这些课程后续会补考的）
        df_one_student_all = df_one_student_all[df_one_student_all['绩点'] != 0]

        # 获取所有的课程号
        df_course_all_ids = df_one_student_all["课程号"]
        course_uniq_ids = set()
        for tId in df_course_all_ids:
            course_uniq_ids.add(tId)

        # 创建用于返回的 Dataframe
        df_ret = DataFrame()

        # 针对每个课程号，选取最大的一行数据
        for tCourseId in course_uniq_ids:
            # 筛选同一课程号下的所有课程
            aaa = df_one_student_all[df_one_student_all["课程号"] == tCourseId]
            # 进行倒叙排列
            bbb = aaa.sort_values(axis=0, ascending=False, by="达成度")
            # 取第一行数据
            ccc = bbb.iloc[0]
            # 获取课程的排序优先级
            course_priority = self.get_course_priority(ccc["课程号"])
            # 将课程信息录入结果DataFrame
            mlst = [ccc["学号"],
                    ccc["姓名"],
                    ccc["开课院系"],
                    ccc["课程号"],
                    ccc["课程名"],
                    ccc["学分"],
                    ccc["达成度"],
                    course_priority]
            myseries = pd.DataFrame([mlst], columns=["学号",
                                                     "姓名",
                                                     "开课院系",
                                                     "课程号",
                                                     "课程名",
                                                     "学分",
                                                     "达成度",
                                                     "优先级"])
            df_ret = pd.concat([df_ret, myseries], ignore_index=True)

        df_ret = df_ret.sort_values(axis=0, ascending=True, by="优先级")
        return df_ret

    def create_final_excel_file(self, stu_id_in, stu_data_in: DataFrame, ):
        """
        生成最终的 Excel 文件
        储存路径为：self.path
        :param stu_id_in: 输入的学生编号
        :param stu_data_in: 输入的学生成绩信息
        :return: 文件绝对路径
        """
        # 保存的excel文件名
        xls_filename = os.path.join(self.out_path,
                                    str(self.gStudentGrade.value),
                                    str(stu_id_in) + ".xlsx")

        # 拷贝生成目标 Excel 文件
        copyfile(template_path + "/template0.xlsx", xls_filename)

        # 打开刚才创建的 workbook
        wb = pyxls.load_workbook(xls_filename)

        # 指定当前显示（活动）的sheet对象
        ws = wb.active

        # 根据年级生成 A1 格的内容
        ws['A1'] = f"计算机科学与技术专业课程目标达成评价（{self.gStudentGrade.value}级）"

        # 生成学号、姓名格内容
        tmp_data = stu_data_in["学号"]
        stu_id = tmp_data[0]
        ws['H3'] = "{} ".format(int(stu_id))

        tmp_data = stu_data_in["姓名"]
        stu_name = tmp_data[0]
        ws['B3'] = stu_name

        # 设置页眉
        ws.HeaderFooter.evenHeader.left.text = f"【学号：{stu_id} / 姓名：{stu_name}】"
        ws.HeaderFooter.oddHeader.left.text = f"【学号：{stu_id} / 姓名：{stu_name}】"

        # dataframe 的行数
        num_df_lines = stu_data_in.shape[0]

        total_cs_course = 0  # 计算机课程计数
        total_other_course = 0  # 其他课程计数
        for tLineIndex in range(num_df_lines):
            # 设置单元格访问符
            if stu_data_in.iloc[tLineIndex, 2] == "计算机系":
                ce_num = "A{}".format(total_cs_course + 7)
                ce_name = "B{}".format(total_cs_course + 7)
                ce_score = "C{}".format(total_cs_course + 7)
                ce_deg = "D{}".format(total_cs_course + 7)
                total_cs_course += 1
            else:
                # 第0列，其他院系独有，开课院系
                ce_college = "F{}".format(total_other_course + 7)
                ws[ce_college] = stu_data_in.iloc[tLineIndex, 2]

                ce_num = "G{}".format(total_other_course + 7)
                ce_name = "H{}".format(total_other_course + 7)
                ce_score = "I{}".format(total_other_course + 7)
                ce_deg = "J{}".format(total_other_course + 7)
                total_other_course += 1

            # 第1列：课程号
            tmp_data = stu_data_in.iloc[tLineIndex, 3]
            ws[ce_num] = tmp_data

            # 第2列：课程名
            tmp_data = stu_data_in.iloc[tLineIndex, 4]
            ws[ce_name] = tmp_data

            # 第3列：学分
            tmp_data = stu_data_in.iloc[tLineIndex, 5]
            ws[ce_score] = tmp_data

            # 第4列：达成度
            tmp_data = stu_data_in.iloc[tLineIndex, 6]
            ws[ce_deg] = tmp_data

        # 计算计算机专业平均达成度，并写入 excel
        ws[f"B{total_cs_course + 7}"] = " 平均达成度："
        tmp_data = "=SUMPRODUCT(D7:D{},C7:C{})/SUM(C7:C{})"
        tmp_data = tmp_data.format(total_cs_course + 6, total_cs_course + 6,
                                   total_cs_course + 6)
        tmp_data = tmp_data if total_cs_course else "无"
        ws[f"D{total_cs_course + 7}"] = tmp_data
        ws[f"D{total_cs_course + 7}"].number_format = "0.000"

        # 计算非计算机专业平均达成度，并存入 excel
        ws[f"H{total_other_course + 7}"] = " 平均达成度："
        tmp_data = "=SUMPRODUCT(I7:I{},J7:J{})/SUM(I7:I{})"
        tmp_data = tmp_data.format(total_other_course + 6, total_other_course + 6,
                                   total_other_course + 6)
        tmp_data = tmp_data if total_other_course else "无"
        ws[f"J{total_other_course + 7}"] = tmp_data
        ws[f"J{total_other_course + 7}"].number_format = "0.000"

        # 设置平均达成度单元格样式
        rows = [total_cs_course + 7]
        cols = ['A', 'B', 'C', 'D']  # 左侧列
        src_rows = [47] * len(rows)
        src_cols = cols
        copy_xlcell_style(ws, rows, cols, src_rows, src_cols)

        rows = [total_other_course + 7]
        cols = ['F', 'G', 'H', 'I', 'J']  # 右侧列
        src_rows = [39] * len(rows)
        src_cols = cols
        copy_xlcell_style(ws, rows, cols, src_rows, src_cols)

        # 设置普通表格单元格样式
        rows = range(7, total_cs_course + 7)
        cols = ['A', 'B', 'C', 'D']  # 左侧列
        src_rows = [7] * len(rows)
        src_cols = cols
        copy_xlcell_style(ws, rows, cols, src_rows, src_cols)

        rows = range(7, total_other_course + 7)
        cols = ['F', 'G', 'H', 'I', 'J']  # 右侧列
        src_rows = [7] * len(rows)
        src_cols = cols
        copy_xlcell_style(ws, rows, cols, src_rows, src_cols)

        # 设置空白单元格样式
        ws[f"A{total_cs_course + 7}"] = ""
        ws[f"C{total_cs_course + 7}"] = ""
        ws[f"F{total_other_course + 7}"] = ""
        ws[f"G{total_other_course + 7}"] = ""
        ws[f"I{total_other_course + 7}"] = ""

        if total_cs_course < 40:
            rows = range(total_cs_course + 8, 48)
            cols = ['A', 'B', 'C', 'D']  # 左侧列
            src_rows = [7] * len(rows)
            src_cols = ['K'] * len(cols)
            copy_xlcell_style(ws, rows, cols, src_rows, src_cols)
            for row in rows:
                for col in cols:
                    ws[f"{col}{row}"] = ""

        if total_other_course < 40:
            rows = range(total_other_course + 8, 48)
            cols = ['F', 'G', 'H', 'I', 'J']  # 右侧列
            src_rows = [7] * len(rows)
            src_cols = ['K'] * len(cols)
            copy_xlcell_style(ws, rows, cols, src_rows, src_cols)
            for row in rows:
                for col in cols:
                    ws[f"{col}{row}"] = ""

        # 保存excel
        wb.save(xls_filename)
        return xls_filename

    def get_file(self, stu_id_in) -> str:
        """
        根据学号生成对应学生的excel文件
        :param stu_id_in: 学号
        :return: excel文件的绝对路径
        """
        # 检测学号输入是否符合规范
        if len(stu_id_in) != 12 or not str(stu_id_in).isdigit():
            raise FormatError("The student id must be a 12-digit number")
        # 修改年级
        if int(str(stu_id_in)[:4]) != self.gStudentGrade.value:
            self.set_garde(EGrads(int(str(stu_id_in)[:4])))
        # 获取该生的所有考试信息
        df_one_student = self.get_one_student_info(stu_id_in)
        # 生成此学生最终的 Excel 文件
        return self.create_final_excel_file(stu_id_in, df_one_student)

    def get_all_file(self, verbose=False):
        """
        生成所有学生的excel文件
        :return: None
        """
        stu_uniq_ids = self.get_student_ids()
        cnt = 0
        for tOneId in stu_uniq_ids:
            file_path = self.get_file(tOneId)
            if verbose:
                print(
                    f"------------------------------------------------------------\n" +
                    f"    num: {cnt}\n" +
                    f"    student Id: {tOneId}\n" +
                    f"    file path: {file_path}\n"
                )


def clear_out_directory():
    """清空输出目录中的所有文件，并创建需要的文件夹"""
    path = out_path + "\\excel"
    for root, dirs, files in os.walk(path, topdown=False):
        for name in files:
            os.remove(os.path.join(root, name))
        for name in dirs:
            os.removedirs(os.path.join(root, name))
    if os.path.exists(path):
        os.removedirs(path)
    os.mkdir(path)
    for i in range(2014, 2017):
        os.mkdir(os.path.join(path, str(i)))


if "__main__" == __name__:
    clear_out_directory()
    for grade in range(2014, 2017):
        # 创建对象实例，并设置要操作的年级编号
        excel_builder: ExcelBuilder = ExcelBuilder(EGrads(grade))
        # 生成所有学生的 excel 文件
        excel_builder.get_all_file(verbose=True)
