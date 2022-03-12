import os
import pickletools
from shutil import copyfile
import pymysql

# 使用 openpyxl 进行读操作
import openpyxl as pyxls
from openpyxl import load_workbook

# 使用 xlsxwriter 进行写操作
import xlsxwriter

# 使用 pandas 处理数据
import pandas as pd
from pandas.core.frame import DataFrame, Series

from config import *
from utils.error import *
from utils.excel import *
from utils.utils import *
from db import db


class DetailProcessor:
    """
    对上传的成绩明细 Excel 文件进行分析
    把成绩数据上传到数据库中
    """
    _template_file_path = os.path.join(temp_dir, "detail_template.xlsx")
    _actual_file_path = os.path.join(temp_dir, "detail_score.xlsx")

    def __init__(self):
        # 输出文件路径
        self.template_file_path: str = DetailProcessor._template_file_path
        self.score_file_path: str = DetailProcessor._actual_file_path
        # 储存接收到的文件
        self.receive_file = None
        # 接收到的文件解析出的 Excel 数据
        self.data_frame: DataFrame = DataFrame()
        # 储存基本课程信息
        self.course_id: str = ""
        self.semester: str = ""
        self.audit_id: int = 0
        # 储存常用数据
        self.form_structure: dict = {}
        self.header: list[str] = []
        self.tot_score: list[int] = []
        self.col_item_id_map: dict[int:int] = {}
        self.template_data: list[list[any]] = []
        self.score_data: list[list[float]] = []

    def __del__(self):
        if os.path.dirname(self.template_file_path) == temp_dir and \
                os.path.exists(self.template_file_path):
            os.remove(self.template_file_path)
        if os.path.dirname(self.score_file_path) == temp_dir and \
                os.path.exists(self.score_file_path):
            os.remove(self.score_file_path)

    def __repr__(self):
        return f"{type(self).__name__}(semester={self.semester}" \
               f", course_id={self.course_id})"

    def set_template_file_path(self, file_path: str):
        self.template_file_path = file_path

    def set_score_file_path(self, file_path: str):
        self.score_file_path = file_path

    def set_receive_file(self, receive_file):
        """
        储存接收到的 FIleLike Object
        :param receive_file:
        :return:
        """
        self.receive_file = receive_file
        # 读取 Excel 文件
        self.data_frame = pd.read_excel(self.receive_file, sheet_name=0)

    def get_course_info(self):
        """
        补全课程信息的课程学期和课程号
        如果存在 audit id 则只根据 audit id 获取信息
        :return:
        """
        if not self.audit_id:
            self.audit_id = db.get_identifier_by_course_info(self.semester,
                                                             self.course_id)
        else:
            self.semester, self.course_id = db.get_course_info_by_identifier(
                self.audit_id)

        if self.audit_id:
            self.analysis_form_structure()

    def set_semester(self, semester: str):
        """
        设置课程学期，同时尝试补全其他信息
        :param semester:
        :return:
        """
        self.semester = semester
        self.get_course_info()

    def set_course_id(self, course_id: str):
        """
        设置课程号，同时尝试补全其他信息
        :param course_id:
        :return:
        """
        self.course_id = course_id
        self.get_course_info()

    def set_audit_id(self, audit_id: int):
        """
        设置 audit id, 并补足其他信息
        :param audit_id:
        :return:
        """
        self.audit_id = audit_id
        self.get_course_info()

    shard_header: list[str] = ["序号", "学号", "姓名", "总评"]

    def analysis_form_structure(self):
        """
        根据数据库中存储的表结构，生成表头，各表头对应的总分，各表头对应的item_id等信息
        生成 self.header
        生成 self.tot_score
        生成 self.col_item_id_map
        :return: 成绩明细表的表头，如果表头中应该是空的，则保存空字符串
        """
        # 检测是否已经获取表结构信息
        if len(self.form_structure) == 0:
            self.form_structure = db.get_form_structure_by_audit_di(self.audit_id)
        header = []
        tot_score = []
        module_sum_cols = []
        col_item_id_map = {}
        now_col = 3  # 记录当前是在那一列，因为前面有三列基本信息
        for module_name, part_info in self.form_structure.items():
            # part_info 直接就是 item_id, 该模组没有 part
            if type(part_info) is int:
                # 直接在 header 中加入 module_name
                header.append(module_name)
                # 同时加入该列的总分信息
                tot_score.append(int(db.get_module_info_by_item_id(part_info)[1]))
                # 记录 column：item_id 的映射
                col_item_id_map[now_col] = part_info
                now_col += 1
                module_sum_cols.append(now_col)
                continue
            t_header = []
            t_tot_score = []
            part_sum_cols = []
            for part_name, item_info in part_info.items():
                # item_info 直接就是 item_id, 该单元没有 item
                if type(item_info) is int:
                    # 直接在 header 中加入 part_name
                    t_header.append(part_name)
                    # 同时加入该列的总分信息，每个单元的总分必是 100
                    t_tot_score.append(100)
                    # 记录 column：item_id 的映射
                    col_item_id_map[now_col] = item_info
                    now_col += 1
                    part_sum_cols.append(now_col)
                    continue
                tt_header = []
                tt_tot_score = []
                for item_name, item_id in item_info.items():
                    # 加入 item 信息列
                    tt_header.append(item_name)
                    # 加入该列的总分信息
                    tt_tot_score.append(int(db.get_item_info_by_item_id(item_id)[1]))
                    # 记录 column：item_id 的映射
                    col_item_id_map[now_col] = item_id
                    now_col += 1
                if len(tt_header) == 1:
                    # 该 part 只有一个 item, 使用 part_name 的名字
                    t_header.append(part_name)
                    part_sum_cols.append(now_col)
                else:
                    # 该 part 有多个 item, 加入一列 part_name
                    tt_header.append(part_name)
                    t_header += tt_header
                    # 将当前记录的总分信息合入总的，因为 part 总分是所有 item 的和，因此加入公式
                    t_tot_score += tt_tot_score
                    col_beg = dec_to_excel_col(now_col - len(tt_tot_score) + 1)
                    col_end = dec_to_excel_col(now_col)
                    t_tot_score.append(f"=SUM({col_beg}%ROW%:{col_end}%ROW%)")
                    part_sum_cols.append(now_col + 1)  # 记录当前所有的part总分的列号
                    # 记录 column：item_id 的映射，此时因为总分信息不用存入数据库，所以置 -1
                    col_item_id_map[now_col] = -1
                    now_col += 1
            if len(t_header) == 1:
                # 此时，这个 module 只有一个 part 并且它只有一个 item, 使用 module_name 覆盖
                header.append(module_name)
                # 简单合并总分信息即可
                tot_score += t_tot_score
            else:
                # 整个 module 有多个 item 直接简单合并即可
                header += t_header
                tot_score += t_tot_score
            if len(part_info) == 1:
                # 此时，这个 module 只有一个 part, 把 part_name 用 module_name 覆盖
                header[-1] = module_name
                module_sum_cols.append(now_col)
            else:
                # 此时，这个 module 有多个 part, 添加 module 总成绩列
                header.append(module_name)
                # 此时记录该列的总分，因为 module 的满分一定是 100, 因此直接加入 100
                col = now_col - 1
                while col_item_id_map[col] == -1:
                    col -= 1
                ratios = [int(r[1]) for r in
                          db.get_parts_info_by_item_id(col_item_id_map[col])]
                ratios_sum = sum(ratios)
                ratios = [i / ratios_sum for i in ratios]
                cells = [f"{dec_to_excel_col(part_sum_cols[i])}%ROW%*{ratios[i]:.3f}"
                         for i in range(len(ratios))]
                func = "+".join(cells)
                tot_score.append(f"={func}")
                module_sum_cols.append(now_col + 1)
                # 记录 column：item_id 的映射，此时因为总分信息不用存入数据库，所以置 -1
                col_item_id_map[now_col] = -1
                now_col += 1
            header.append("")
            tot_score.append("")
            # 依旧加入 -1, 表示此列数据不用存入数据库
            col_item_id_map[now_col] = -1
            now_col += 1
        # 加入总评列的公式
        ratios = [info[1] for info in db.get_modules_info_by_audit_id(self.audit_id)]
        module_sum_cols = [dec_to_excel_col(col) for col in module_sum_cols]
        func = "+".join([f"{module_sum_cols[i]}%ROW%*{ratios[i] / 100}"
                         for i in range(len(ratios))])
        tot_score.append(f"={func}")

        self.header = DetailProcessor.shard_header[:3] + header \
                      + DetailProcessor.shard_header[3:]
        self.tot_score = tot_score
        # 设置前3列和最后一列的 item_id 为 -1
        col_item_id_map[0] = col_item_id_map[1] = col_item_id_map[2] = -1
        col_item_id_map[now_col] = -1
        self.col_item_id_map = col_item_id_map
        # 储存模板的数据
        self.template_data = [
            [1, 201910311111, "张三", *tot_score],
            [2, 201910322222, "李四", *tot_score],
            [3, 201910333333, "王五", *tot_score]
        ]

    def write_header(self, work_sheet: xlsxwriter.workbook.Worksheet, formats: dict):
        # 设置表头与列宽
        for i, val in enumerate(self.header):
            # 设置表头
            if val:
                work_sheet.write_string(0, i, val, formats["header_zh"])
            # 设置列宽
            if val == "序号":
                work_sheet.set_column(i, i, 6)
            if val == "学号":
                work_sheet.set_column(i, i, 13)
            if not val:
                work_sheet.set_column(i, i, 2)

    def create_final_template(self, template_file_path=_template_file_path):
        """
        根据 file_path 生成课程成绩明细模板表
        :param template_file_path: 课程成绩明细模板表的输出位置
        :return:
        """
        # 保存输出文件目录
        self.set_template_file_path(template_file_path)
        # 生成表头，模板预制数据，和列:item_id映射信息
        if not self.header or not self.template_data:
            self.analysis_form_structure()
        # 新建 xlsx 文件
        work_book: xlsxwriter.Workbook = xlsxwriter.Workbook(template_file_path)
        # 新建 工作表
        work_sheet: xlsxwriter.workbook.Worksheet = work_book.add_worksheet()

        # 在表中插入样式
        formats = insert_format(work_book)

        # 写入表头
        self.write_header(work_sheet, formats)

        # 写入数据样例
        write_data(work_sheet, formats, DataFrame(self.template_data))

        work_book.close()

    def get_template(self):
        """
        生成储存在临时文件夹中的模板表
        :return: 生成的文件的路径
        """
        # 生成文件名
        timestamp = get_timestamp().replace(' ', '_')
        file_path = os.path.join(temp_dir, f"detail_template_{timestamp}.xlsx")
        self.create_final_template(file_path)
        return file_path

    def read_data(self):
        # 读取成绩明细 Excel 信息
        self.data_frame = pd.read_excel(self.receive_file, sheet_name=0)
        # 生成列:item_id映射
        if not self.template_data:
            self.analysis_form_structure()

        # 清理数据，将表头为空的列的名字置 ""
        for col_name in self.data_frame.columns:
            if col_name not in self.header:
                self.data_frame = self.data_frame.rename(columns={col_name: ""})

    def upload_data(self):
        if self.data_frame.empty:
            raise ValueError("Excel file is empty.")

        self.read_data()  # 读取数据

        stu_ids: Series = self.data_frame["学号"].astype("int64")
        stu_name: Series = self.data_frame["姓名"].astype("string")

        # 保存学生基本信息
        infos = []
        for i in range(len(stu_ids)):
            infos.append((stu_ids.iloc[i], stu_name.iloc[i],
                          self.semester, self.course_id))
        db.save_student_infos(infos)
        # 保存所有成绩信息
        infos = []
        for j in range(self.data_frame.shape[1]):
            if self.col_item_id_map[j] == -1:
                continue
            scores: Series = self.data_frame.iloc[:, j].astype(float)
            for i in range(self.data_frame.shape[0]):
                infos.append((self.col_item_id_map[j], stu_ids.iloc[i], scores.iloc[i]))
        db.save_item_scores(infos)

    def get_score_data(self) -> DataFrame:
        res = DataFrame(columns=["序号", "学号", "姓名"])

        # 获取学生基本信息
        stu_info = db.get_students_info_by_audit_id(self.audit_id)
        for i, row in enumerate(stu_info):
            df = DataFrame(data={"序号": [i + 1], "学号": [int(row[0])], "姓名": [row[1]]})
            res = pd.concat([res, df], ignore_index=True)

        # 获取数据库中成绩数据
        for i in range(3, len(self.col_item_id_map)):
            item_id = self.col_item_id_map[i]
            if self.header[i] == "":
                df = DataFrame(data=[""] * res.shape[0], columns=[""])
                res = pd.concat([res, df], axis=1)
            elif item_id == -1:
                # 加入对应的公式
                df = DataFrame(
                    data={
                        self.header[i]: [
                            self.template_data[0][i].replace("%ROW%", str(j + 2))
                            for j in range(res.shape[0])]
                    }, columns=[self.header[i]]
                )
                res = pd.concat([res, df], axis=1)
            else:
                # 记录成绩
                data = db.get_score_info_by_item_id(item_id)
                df = DataFrame(
                    data={"学号": [int(row[0]) for row in data],
                          self.header[i]: [int(row[1]) for row in data]},
                    columns=["学号", self.header[i]]
                )
                res = res.merge(df, how="left", left_on="学号", right_on="学号")

        return res

    def create_final_score_file(self, score_file_path: str = _actual_file_path):
        # 保存实际成绩文件输出目录
        self.set_score_file_path(score_file_path)

        # 获取成绩数据
        data = self.get_score_data()

        # 新建 xlsx 文件
        work_book: xlsxwriter.Workbook = xlsxwriter.Workbook(score_file_path)
        # 新建 工作表
        work_sheet: xlsxwriter.workbook.Worksheet = work_book.add_worksheet()

        # 在表中插入样式
        formats = insert_format(work_book)

        # 写入表头
        self.write_header(work_sheet, formats)

        # 写入数据
        write_data(work_sheet, formats, data)

        work_book.close()

        return score_file_path

    def get_score_file(self):
        # 为文件打上时间戳，避免访问冲突
        timestamp = get_timestamp().replace(' ', '_')
        path = os.path.join(temp_dir, f"score_detail_{timestamp}.docx")
        return self.create_final_score_file(path)


if __name__ == '__main__':
    processor: DetailProcessor = DetailProcessor()
    processor.set_receive_file(open(
        os.path.join(os.path.dirname(__file__), "test\\detail_pre.xlsx"),
        "rb"))
    processor.set_semester("2023-2024-2")
    processor.set_course_id("XX110110")
    processor.upload_data()
    processor.create_final_template(
        os.path.join(os.path.dirname(__file__), "test\\output_template.xlsx"))
    processor.create_final_score_file(
        os.path.join(os.path.dirname(__file__), "test\\output_score.xlsx"))
    input("Press Enter to exit...")
